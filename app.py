from flask import Flask, render_template, jsonify
import requests
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)

# 🔹 Replace with your OneDrive direct download link
EXCEL_URL = "https://onedrive.live.com/download?resid=11FE308CD3290E40/IQSxIFf7Fl0CTY6XKxciStPoAflVanrHAEW_aVs5SMSkygM&download=1"
  # example format

def get_month_data(month_name):
    try:
        r = requests.get(EXCEL_URL)
        r.raise_for_status()  # check for download errors
    except requests.exceptions.RequestException as e:
        return {"error": f"Cannot fetch Excel file: {e}"}

    file_stream = BytesIO(r.content)
    wb = load_workbook(file_stream)
    sheet = wb.active

    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        month, paid, days_in_month, days_absent, days_coming, amount = row[:6]
        if month and month.strip().upper() == month_name.upper():
            data = {
                "Month": month,
                "Paid": paid,
                "Days in Month": days_in_month,
                "Days Absent": days_absent,
                "Days Coming": days_coming,
                "Amount": amount
            }
            break
    wb.close()
    return data

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/month/<month_name>")
def month_data(month_name):
    data = get_month_data(month_name)
    if data:
        return jsonify(data)
    else:
        return jsonify({"error": "No data found for this month"})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)


